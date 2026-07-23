#!/usr/bin/env python3
"""Liest das (aus Google Sheet exportierte) .xlsx und schreibt data/inhalte.json.

Nur die im Sheet gepflegten Spalten werden übernommen. Die nicht im Sheet
sichtbare "Garnitur" der Stellen (Zusatz-Badge, Wizard-Untertitel, langer
Google-for-Jobs-Text, employmentType, Slug) bleibt für bestehende Stellen
erhalten; ändert der Kunde die Beschreibung, wird der Google-Text neu erzeugt;
für neue Stellen wird ein Slug erzeugt und die Garnitur später automatisch
aus Defaults ergänzt (siehe apply-content.mjs).

Aufruf:  python3 _generatoren/merge-sheet.py <export.xlsx>
"""
import json, sys, re, unicodedata, os
from openpyxl import load_workbook

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..")) + "/"
xlsx = sys.argv[1] if len(sys.argv) > 1 else ROOT + "data/sheet-export.xlsx"

old = json.load(open(ROOT + "data/inhalte.json"))
old_jobs = {j["titel_de"]: j for j in old.get("stellen", [])}
old_slugs = {j["stelle"] for j in old.get("stellen", [])}

wb = load_workbook(xlsx, data_only=True)

def rows(ws):
    """Zeilen ab 2 als Liste; Kopf­zeile bestimmt die Reihenfolge."""
    it = ws.iter_rows(values_only=True)
    next(it, None)  # Kopfzeile überspringen
    return [r for r in it if any(c not in (None, "") for c in r)]

def cell(r, i):
    return "" if i >= len(r) or r[i] is None else str(r[i]).strip()

def slugify(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = re.sub(r"\(.*?\)", "", s)                 # (m/w/d) entfernen
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return s or "stelle"

# ---------- Stellenangebote ----------
# Spalten: Status | Stellentitel (DE) | Job title (EN) | Beschäftigung | Beschreibung (DE) | Description (EN)
stellen = []
used = set()
for r in rows(wb["Stellenangebote"]):
    titel_de = cell(r, 1)
    if not titel_de:
        continue
    beschr_de = cell(r, 4)
    prev = old_jobs.get(titel_de, {})
    slug = prev.get("stelle")
    if not slug:                                   # neue Stelle → eindeutigen Slug bilden
        base = slugify(titel_de); slug = base; n = 2
        while slug in old_slugs or slug in used:
            slug = f"{base}-{n}"; n += 1
    used.add(slug)
    job = {
        "stelle": slug,
        "titel_de": titel_de,
        "titel_en": cell(r, 2),
        "beschaeftigung": cell(r, 3) or "Vollzeit",
        "beschreibung_de": beschr_de,
        "beschreibung_en": cell(r, 5),
        # Garnitur aus Bestand übernehmen:
        "zusatz_de": prev.get("zusatz_de", ""), "zusatz_en": prev.get("zusatz_en", ""),
        "sub_de": prev.get("sub_de", ""), "sub_en": prev.get("sub_en", ""),
        "jpTitle": prev.get("jpTitle", ""),
        "jpEmp": prev.get("jpEmp", ""),
        "jp_de": prev.get("jp_de", ""),
        "status": (cell(r, 0) or "offen").lower(),
    }
    # Beschreibung geändert (oder neu) → Google-for-Jobs-Text neu erzeugen lassen
    if beschr_de != prev.get("beschreibung_de", ""):
        job["jp_de"] = ""
    stellen.append(job)

# ---------- Referenzen ----------
# Status|Slug|Branche|Bild-Dateiname|Titel(DE)|Title(EN)|Kurztext(DE)|Short(EN)|
# Aufgabe(DE)|Task(EN)|Lösung(DE)|Solution(EN)|Eckdaten(DE)|Key data(EN)|Leistungen
old_refs = {x["slug"]: x for x in old.get("referenzen", [])}
referenzen = []
for r in rows(wb["Referenzen"]):
    slug = cell(r, 1)
    if not slug:
        continue
    prev = old_refs.get(slug, {})
    referenzen.append({
        "slug": slug,
        "branche": cell(r, 2) or "lebensmittel",
        "bild": cell(r, 3),
        "bildW": prev.get("bildW", ""), "bildH": prev.get("bildH", ""),  # wird eh aus Datei gelesen
        "titel_de": cell(r, 4), "titel_en": cell(r, 5),
        # Bild-Alt-Text ist nicht im Sheet → aus Bestand halten; bei neuen Referenzen
        # greift in apply-content der Fallback auf den Titel.
        "alt_de": prev.get("alt_de", ""), "alt_en": prev.get("alt_en", ""),
        "kurz_de": cell(r, 6), "kurz_en": cell(r, 7),
        "aufgabe_de": cell(r, 8), "aufgabe_en": cell(r, 9),
        "loesung_de": cell(r, 10), "loesung_en": cell(r, 11),
        "eck_de": cell(r, 12), "eck_en": cell(r, 13),
        "leistungen": cell(r, 14),
        "status": (cell(r, 0) or "sichtbar").lower(),
    })

out = {"referenzen": referenzen, "stellen": stellen}
json.dump(out, open(ROOT + "data/inhalte.json", "w"), ensure_ascii=False, indent=2)
print(f"✓ merge-sheet: {len(referenzen)} Referenzen, {len(stellen)} Stellen → data/inhalte.json")
print("  Stellen:", " | ".join(f"{j['stelle']}({j['status']})" for j in stellen))
print("  Versteckte Referenzen:", ", ".join(x["slug"] for x in referenzen if x["status"] == "versteckt") or "keine")
